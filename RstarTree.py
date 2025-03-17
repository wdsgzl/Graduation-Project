import openpyxl
import sys
sys.setrecursionlimit(1000000000)
class Cuboid:
    def __init__(self, x1, y1, z1, x2, y2, z2):
        self.x1, self.y1, self.z1 = x1, y1, z1  # 左下前角坐标
        self.x2, self.y2, self.z2 = x2, y2, z2  # 右后上角坐标

    def volume(self):
        """计算长方体的体积"""
        return abs(self.x2 - self.x1) * abs(self.y2 - self.y1) * abs(self.z2 - self.z1)

    def overlap(self, other):
        """计算与另一个长方体的重叠体积"""
        x_overlap = max(0, abs(min(self.x2, other.x2) - max(self.x1, other.x1)))
        y_overlap = max(0, abs(min(self.y2, other.y2) - max(self.y1, other.y1)))
        z_overlap = max(0, abs(min(self.z2, other.z2) - max(self.z1, other.z1)))
        if x_overlap==0 and y_overlap==0 and z_overlap==0:
            return 0
        else:
            return x_overlap * y_overlap * z_overlap

    def merge(self, other):
        """合并两个长方体，生成包含二者的最小边界长方体（MBC）"""
        x1 = min(self.x1, other.x1)
        y1 = min(self.y1, other.y1)
        z1 = min(self.z1, other.z1)
        x2 = max(self.x2, other.x2)
        y2 = max(self.y2, other.y2)
        z2 = max(self.z2, other.z2)
        return Cuboid(x1, y1, z1, x2, y2, z2)

    def __repr__(self):
        return f"Cuboid({self.x1}, {self.y1}, {self.z1}, {self.x2}, {self.y2}, {self.z2})"

class R3StarTreeNode:
    def __init__(self, is_leaf=False):
        self.is_leaf = is_leaf        # 是否为叶子节点
        self.cuboids = []             # 节点中的长方体（MBC）
        self.children = []            # 子节点（如果是非叶子节点）
        self.parent = None            # 父节点

    def compute_mbc(self):
        """计算节点的最小边界长方体（MBC）"""
        if not self.cuboids:
            return None
        x1 = min(c.x1 for c in self.cuboids)
        y1 = min(c.y1 for c in self.cuboids)
        z1 = min(c.z1 for c in self.cuboids)
        x2 = max(c.x2 for c in self.cuboids)
        y2 = max(c.y2 for c in self.cuboids)
        z2 = max(c.z2 for c in self.cuboids)
        return Cuboid(x1, y1, z1, x2, y2, z2)

    def add_child(self, child):
        """添加子节点并更新父子关系"""
        self.children.append(child)
        child.parent = self
        self.cuboids.append(child.compute_mbc())  # 更新当前节点的MBC

    def __repr__(self):
        return f"Node(Leaf={self.is_leaf}, MBCs={self.cuboids})"


class R3StarTree:
    def __init__(self, max_entries=4, min_entries=2):
        self.max_entries = max_entries  # 节点最大条目数
        self.min_entries = min_entries  # 节点最小条目数
        self.root = R3StarTreeNode(is_leaf=True)  # 根节点初始为叶子

    def insert(self, cuboid):
        """插入一个长方体到R*树"""
        leaf = self._choose_leaf(self.root, cuboid)
        leaf.cuboids.append(cuboid)

        # 如果节点溢出，进行分裂或重新插入
        if len(leaf.cuboids) > self.max_entries:
            self._handle_overflow(leaf)

    def _choose_leaf(self, node, cuboid):
        """迭代选择插入的叶子节点（基于最小体积扩展）"""
        current_node = node
        while True:
            if current_node.is_leaf:
                return current_node

            min_volume_increase = float('inf')
            best_child = None

            # 遍历子节点，选择体积扩展最小的子节点
            for child in current_node.children:
                mbc = child.compute_mbc()
                merged_mbc = mbc.merge(cuboid) if mbc else cuboid
                volume_increase = merged_mbc.volume() - (mbc.volume() if mbc else 0)

                if volume_increase < min_volume_increase:
                    min_volume_increase = volume_increase
                    best_child = child

            if best_child is None:
                break  # 没有找到子节点，返回当前节点
            else:
                current_node = best_child

        return current_node

    def _handle_overflow(self, node):
        """处理节点溢出（优先重新插入，否则分裂）"""
        # 强制重新插入部分条目（优化策略）
        entries_to_reinsert = node.cuboids[self.max_entries - self.min_entries:]
        node.cuboids = node.cuboids[:self.min_entries]

        # 直接分裂节点，避免重新插入导致递归
        self._split_node(node)

    def _split_node(self, node):
        """三维分裂策略（迭代处理父节点溢出）"""
        while True:
            # --- 原分裂逻辑（计算最佳分裂轴）---
            best_axis = None
            best_cost = float('inf')
            for axis in ['x', 'y', 'z']:
                sorted_entries = sorted(node.cuboids, key=lambda c: getattr(c, f"{axis}1"))
                split_index = len(sorted_entries) // 2
                group1 = sorted_entries[:split_index]
                group2 = sorted_entries[split_index:]
                mbc1 = self._merge_all(group1)
                mbc2 = self._merge_all(group2)
                total_volume = mbc1.volume() + mbc2.volume()
                if total_volume < best_cost:
                    best_cost = total_volume
                    best_axis = axis

            # --- 执行分裂 ---
            sorted_entries = sorted(node.cuboids, key=lambda c: getattr(c, f"{best_axis}1"))
            split_index = len(sorted_entries) // 2

            new_node = R3StarTreeNode(is_leaf=node.is_leaf)
            new_node.cuboids = sorted_entries[split_index:]
            new_node.children = node.children[split_index:] if not node.is_leaf else []

            node.cuboids = sorted_entries[:split_index]
            node.children = node.children[:split_index] if not node.is_leaf else []

            # --- 更新父节点（迭代处理父节点溢出）---
            if node.parent is None:
                # 根节点分裂，创建新根
                new_root = R3StarTreeNode()
                new_root.add_child(node)
                new_root.add_child(new_node)
                self.root = new_root
                break
            else:
                # 将新节点添加到父节点
                parent = node.parent
                parent.add_child(new_node)

                # 检查父节点是否溢出，若溢出则继续分裂
                if len(parent.cuboids) <= self.max_entries:
                    break
                else:
                    node = parent  # 向上处理父节点
    def _merge_all(self, cuboids):
        """合并一组长方体为一个MBC"""
        if not cuboids:
            return None
        x1 = min(c.x1 for c in cuboids)
        y1 = min(c.y1 for c in cuboids)
        z1 = min(c.z1 for c in cuboids)
        x2 = max(c.x2 for c in cuboids)
        y2 = max(c.y2 for c in cuboids)
        z2 = max(c.z2 for c in cuboids)
        return Cuboid(x1, y1, z1, x2, y2, z2)

    def search(self, query_cuboid):
        """查询与给定长方体相交的所有条目"""
        results = []

        self._search_recursive(self.root, query_cuboid, results)
        return results

    def _search_recursive(self, node, query_cuboid, results):
        """递归搜索"""
        print(f"当前节点: {node}")
        if node.is_leaf:
            print("叶子节点，检查长方体...")
            for cuboid in node.cuboids:
                print(f"检查长方体: {cuboid}, 重叠体积: {cuboid.overlap(query_cuboid)}")
                if cuboid.overlap(query_cuboid) >= 0:
                    results.append(cuboid)
        else:
            print("非叶子节点，递归搜索子节点...")
            for child in node.children:
                mbc = child.compute_mbc()
                print(f"子节点 MBC: {mbc}, 重叠体积: {mbc.overlap(query_cuboid)}")
                if mbc.overlap(query_cuboid) >= 0:
                    self._search_recursive(child, query_cuboid, results)

    def print_tree(self, node=None, level=0):
        """打印树结构"""
        if node is None:
            node = self.root
        print(f"Level {level}: {node}")
        if not node.is_leaf:
            for child in node.children:
                self.print_tree(child, level + 1)
def getdata(file_path):
    colum=[1,2,4]
    bk = openpyxl.load_workbook(file_path)
    sheet=bk.active
    minrow = sheet.min_row  # 最小行
    maxrow = sheet.max_row  # 最大行
    global x
    x=[]
    global y
    y=[]
    global z
    z=[]
    for m in colum:#0,1,3
        if(m==1):
            for n in range(minrow, maxrow+1):
                x.append(sheet.cell(n, m).value)
                #print(x, end=" ")
        if (m == 2):
            for n in range(minrow, maxrow + 1):
                y.append(sheet.cell(n, m).value)
                #print(y, end=" ")
        if (m == 4):
            for n in range(minrow, maxrow + 1):
                z.append(sheet.cell(n, m).value)
                #print(z, end=" ")
    #print(x)

if __name__ == "__main__":
    file_path='data.xlsx'
    getdata(file_path)

    # 创建三维R*树
    r3star_tree = R3StarTree(max_entries=4, min_entries=2)

    # 插入三维长方体
    cuboids = []
    for i in range(len(x) - 1):
        if i < len(x) - 1:
            x1 = x[i]
            x2 = x[i + 1]
            y1 = y[i]
            y2 = y[i + 1]
            z1 = z[i]
            z2 = z[i + 1]
        cuboids.append(Cuboid(float(x1), float(y1), float(z1), float(x2), float(y2), float(z2)))
    for c in cuboids:
        r3star_tree.insert(c)

    # 打印树结构
    print("三维R*树结构:")
    r3star_tree.print_tree()

    # 查询与 [0.5, 0.5, 0.5] 到 [2.5, 2.5, 2.5] 相交的长方体
    query = Cuboid(0,0,0,1000,1000,1000)
    results = r3star_tree.search(query)
    print(f"查询结果: {results}")